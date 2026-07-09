"""
Exporta os dados atualmente cadastrados (antecedentes, talentos, armas,
equipamentos, criaturas, pericias, titulos de sobrevivencia e regras base)
para uma planilha Excel, para revisao antes de ajustar o banco / construir
a ficha interativa.

Le sempre direto dos .docx (via os servicos de pagina), nao do zona_z.db
em cache, entao sempre reflete o conteudo mais recente.

Uso:
    python export_to_excel.py
"""
import sys
import os

sys.path.insert(0, os.path.dirname(__file__))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.pagina3_service import get_pagina3_data
from app.services.pagina4_service import get_pagina4_data
from app.services.pagina6_service import get_pagina6_data
from app.api.character_builder import PERICIAS_CATEGORIAS, TITULOS_SOBREVIVENCIA, REGRAS_BASE

OUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'Database', 'Zona-Z_Banco_Dados.xlsx')

HEADER_FILL = PatternFill('solid', fgColor='8B0000')
HEADER_FONT = Font(color='FFFFFF', bold=True)


def _join(value):
    if isinstance(value, list):
        return ' | '.join(str(v) for v in value if str(v).strip())
    return value if value is not None else ''


def _write_sheet(wb, name, columns, rows):
    ws = wb.create_sheet(name)
    ws.append([label for _, label in columns])
    for c in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row in rows:
        ws.append([_join(row.get(key, '')) for key, _ in columns])
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for c, (key, label) in enumerate(columns, start=1):
        longest = max((len(str(_join(r.get(key, '')))) for r in rows), default=0)
        ws.column_dimensions[get_column_letter(c)].width = max(10, min(60, longest + 2, len(label) + 4))
    print(f'  {name}: {len(rows)} linhas')
    return ws


# ---------------------------------------------------------------- Antecedentes

def build_antecedentes():
    data = get_pagina3_data()
    rows = []
    for a in data.get('antecedentes', []):
        talento = a.get('talento', {})
        rows.append({
            'antecedente': a['titulo'],
            'tracos': a.get('tracos', ''),
            'ocupacoes': a.get('ocupacoes', ''),
            'forcas': a.get('forcas', ''),
            'fraquezas': a.get('fraquezas', ''),
            'bonus': a.get('bonus', []),
            'pacote': a.get('pacote', []),
            'pericia_primaria': a.get('pericia_primaria', ''),
            'pericias_secundarias': a.get('pericias_secundarias', ''),
            'talento_nome': talento.get('nome', ''),
            'talento_tipo': talento.get('tipo', ''),
            'talento_descricao': talento.get('descricao', ''),
            'talento_efeito': talento.get('efeito', ''),
            'descricao': a.get('descricao', []),
        })
    columns = [
        ('antecedente', 'Antecedente'),
        ('tracos', 'Tracos Marcantes'),
        ('ocupacoes', 'Ocupacoes Anteriores'),
        ('forcas', 'Forcas no Novo Mundo'),
        ('fraquezas', 'Fraquezas no Novo Mundo'),
        ('bonus', 'Bonus Inicial'),
        ('pacote', 'Pacote Inicial'),
        ('pericia_primaria', 'Pericia Primaria'),
        ('pericias_secundarias', 'Pericias Secundarias'),
        ('talento_nome', 'Talento Exclusivo'),
        ('talento_tipo', 'Talento - Tipo'),
        ('talento_descricao', 'Talento - Descricao'),
        ('talento_efeito', 'Talento - Efeito'),
        ('descricao', 'Descricao'),
    ]
    return columns, rows


# -------------------------------------------------------------------- Talentos

def build_talentos():
    data = get_pagina3_data()
    rows = []
    for categoria in data.get('talentos', []):
        for item in categoria.get('itens', []):
            rows.append({
                'classe': categoria.get('titulo', ''),
                'subcategoria': item.get('subcategoria', ''),
                'nome': item.get('nome', ''),
                'tipo': item.get('tipo', ''),
                'descricao': item.get('descricao', ''),
                'efeito': item.get('efeito', ''),
                'uso': item.get('uso', ''),
            })
    columns = [
        ('classe', 'Classe (Categoria)'),
        ('subcategoria', 'Subcategoria'),
        ('nome', 'Nome'),
        ('tipo', 'Tipo'),
        ('descricao', 'Descricao'),
        ('efeito', 'Efeito'),
        ('uso', 'Uso'),
    ]
    return columns, rows


# --------------------------------------------------- Armas e Equipamentos (pag. 4)

EQUIP_COLUMNS = [
    ('capitulo', 'Capitulo'),
    ('secao', 'Secao'),
    ('subcategoria', 'Subcategoria'),
    ('titulo_sobrevivencia', 'Titulo de Sobrevivencia'),
    ('nome', 'Nome'),
    ('tipo', 'Tipo'),
    ('categoria', 'Categoria'),
    ('dano_base', 'Dano Base'),
    ('dano', 'Dano'),
    ('modificador', 'Modificador'),
    ('defesa', 'Defesa'),
    ('regiao_corpo', 'Regiao do Corpo'),
    ('alcance', 'Alcance'),
    ('mao_usada', 'Mao Usada'),
    ('carga', 'Carga'),
    ('capacidade', 'Capacidade'),
    ('municao', 'Municao'),
    ('som', 'Som'),
    ('penalidade', 'Penalidade'),
    ('pre_requisito', 'Pre-requisito'),
    ('equipamento', 'Equipamento'),
    ('uso', 'Uso'),
    ('durabilidade', 'Durabilidade'),
    ('efeito', 'Efeito'),
    ('efeito_adicional', 'Efeito Adicional'),
    ('descricao', 'Descricao'),
    ('componentes', 'Componentes'),
    ('materiais', 'Materiais Necessarios'),
    ('itens_iniciais', 'Itens Iniciais'),
    ('combo', 'Combos'),
    ('imagem', 'Imagem (arquivo)'),
]


def _flatten_equip_section(node, capitulo_titulo, path, out):
    for item in node.get('itens', []):
        row = dict(item)
        row['capitulo'] = capitulo_titulo
        row['secao'] = ' > '.join(path)
        out.append(row)
    for sub in node.get('subsecoes', []):
        _flatten_equip_section(sub, capitulo_titulo, path + [sub['titulo']], out)


def build_equipamentos():
    data = get_pagina4_data()
    rows = []
    for cap in data.get('capitulos', []):
        for h3 in cap.get('secoes', []):
            _flatten_equip_section(h3, cap['titulo'], [h3['titulo']], rows)
    return EQUIP_COLUMNS, rows


# ------------------------------------------------------------- Criaturas (pag. 6)

CREATURE_COLUMNS = [
    ('capitulo', 'Capitulo'),
    ('secao', 'Secao'),
    ('nome', 'Nome'),
    ('tipo', 'Tipo'),
    ('ca', 'CA'),
    ('pv', 'PV'),
    ('movimento', 'Movimento'),
    ('ataque', 'Ataque'),
    ('dano', 'Dano'),
    ('critico', 'Critico'),
    ('reducao_de_danos', 'Reducao de Danos'),
    ('poder_especial', 'Poder Especial'),
    ('habilidade_unica', 'Habilidade Unica'),
    ('comportamento', 'Comportamento'),
    ('efeito', 'Efeito'),
    ('municao', 'Municao'),
    ('pericias', 'Pericias'),
    ('dano_locais', 'Dano por Local Atingido'),
    ('observacao', 'Observacao'),
    ('descricao', 'Descricao'),
    ('imagem', 'Imagem (arquivo)'),
]


def _flatten_creature_section(node, capitulo_titulo, path, out):
    for creature in node.get('criaturas', []):
        row = dict(creature)
        row['capitulo'] = capitulo_titulo
        row['secao'] = ' > '.join(path)
        out.append(row)
    for sub in node.get('subsecoes', []):
        _flatten_creature_section(sub, capitulo_titulo, path + [sub['titulo']], out)


def build_criaturas():
    data = get_pagina6_data()
    rows = []
    for cap in data.get('capitulos', []):
        _flatten_creature_section(cap, cap['titulo'], [], rows)
    return CREATURE_COLUMNS, rows


# ------------------------------------------------------- Pericias / Titulos / Regras

def build_pericias():
    rows = []
    for categoria, itens in PERICIAS_CATEGORIAS.items():
        for pericia in itens:
            rows.append({'categoria': categoria, 'pericia': pericia})
    columns = [('categoria', 'Categoria'), ('pericia', 'Pericia')]
    return columns, rows


def build_titulos_sobrevivencia():
    rows = [{'ordem': i + 1, 'titulo': t} for i, t in enumerate(TITULOS_SOBREVIVENCIA)]
    columns = [('ordem', 'Ordem'), ('titulo', 'Titulo de Sobrevivencia')]
    return columns, rows


def build_regras_base():
    rows = [{'regra': k, 'valor': v} for k, v in REGRAS_BASE.items()]
    columns = [('regra', 'Regra'), ('valor', 'Valor')]
    return columns, rows


def main():
    wb = Workbook()
    wb.remove(wb.active)

    print('Gerando planilha...')
    _write_sheet(wb, 'Antecedentes', *build_antecedentes())
    _write_sheet(wb, 'Talentos', *build_talentos())
    _write_sheet(wb, 'Armas e Equipamentos', *build_equipamentos())
    _write_sheet(wb, 'Criaturas', *build_criaturas())
    _write_sheet(wb, 'Pericias', *build_pericias())
    _write_sheet(wb, 'Titulos de Sobrevivencia', *build_titulos_sobrevivencia())
    _write_sheet(wb, 'Regras Base', *build_regras_base())

    wb.save(OUT_PATH)
    print()
    print(f'Planilha salva em: {os.path.abspath(OUT_PATH)}')


if __name__ == '__main__':
    main()
